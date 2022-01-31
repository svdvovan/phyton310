from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from tkinter import *

root = Tk()
root.geometry('800x100+1000+300')

fileName = 'ali_gui.xlsx'

l_url = Label(root, text="Введите url страницы:").grid(row=0, column=0, pady=10)
e_url = Entry(root, width=100)
e_url.grid(row=0, column=1, padx=15, sticky=E)
excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='ali2', index=0)

count = 2
countRow = 2


def bigParser(e_url):
    global count
    global countRow

    excel_sheet['A1'] = 'id'
    excel_sheet['B1'] = 'Название'
    excel_sheet['C1'] = 'Цена'
    excel_sheet['D1'] = 'Категория'
    excel_sheet['E1'] = 'Описание'
    countCol = 6
    countFoto = 1
    CountP = 1
    CountPn = 17
    CountPnZ = 18
    CountPZZ = 1
    largeFoto2 = []

    for url in e_url:

        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')
        name = soup.find('h1').get_text()
        foto = soup.find('div', class_='Product_GalleryBar__bar__156z6').find_all('img')
        try:
            price = soup.find('span',
                              class_='ali-kit_Base__base__1odrub ali-kit_Base__default__1odrub ali-kit_Base__strong__1odrub price ali-kit_Price__size-xl__12ybyf Product_Price__current__1uqb8 product-price-current').getText()
        except:
            price = 100
        finally:
            descriptionRaw = soup.find('div', class_='ProductDescription-module_content__1xpeo')
            description = (str(descriptionRaw)).replace("\n", '')

            svoistva = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
                                                                                                 class_='ali-kit_Base__base__1odrub ali-kit_Base__light__1odrub')

            svoistva_znach = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
                                                                                                       class_='ali-kit_Base__base__1odrub ali-kit_Base__default__1odrub')

            print(description)
            print(price)
            print(name)

            # excel_sheet[f'A{count}'] = id
            excel_sheet[f'B{count}'] = name
            excel_sheet[f'C{count}'] = price
            excel_sheet[f'D{count}'] = "игрушки"
            excel_sheet[f'E{count}'] = description

            for f in foto:
                largeFoto = f['src']
                largeFoto2.append(largeFoto.replace('_50x50.jpg', ''))
            print(largeFoto2)

            for fot in largeFoto2:
                excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
                excel_sheet.cell(row=countRow, column=countCol).value = fot
                countFoto += 1
                countCol += 1

            datas = [i.get_text(strip=True) for i in svoistva]
            datas2 = [i.get_text() for i in svoistva_znach]

            for data in datas:
                excel_sheet.cell(row=1, column=CountPn).value = f'Свойство {CountP}'
                excel_sheet.cell(row=countRow, column=CountPn).value = data.replace(':', '')
                CountP += 1
                CountPn += 2

            for data2 in datas2:
                excel_sheet.cell(row=1, column=CountPnZ).value = f'Значение {CountPZZ}'
                excel_sheet.cell(row=countRow, column=CountPnZ).value = data2
                CountPnZ += 2
                CountPZZ += 1

            count += 1
            countRow += 1

    excel_file.save(fileName)


def getUrl():
    print(e_url.get())
    urls = [e_url.get()]
    bigParser(urls)


l_start = Button(root, text="Парсить", command=getUrl).grid(row=1, column=0, padx=10, pady=10, sticky=W)
l_clear = Button(root, text="Очистить", command=lambda: e_url.delete(0, END)).grid(row=1, column=1, padx=0, pady=10,
                                                                                   sticky=W)
l_close = Button(root, text="Закрыть", command=lambda: root.destroy()).grid(row=1, column=1, padx=20, pady=10, sticky=E)
root.mainloop()
