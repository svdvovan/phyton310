from bs4 import BeautifulSoup
import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook

urls = [
    'https://aliexpress.ru/item/1005001465081562.html?spm=a2g20.search.0.0.5d9f39d7iO6QQb&algo_pvid=5d1dd444-7edd-4673-b275-d97a298ad63a&algo_expid=5d1dd444-7edd-4673-b275-d97a298ad63a-12',
    'https://aliexpress.ru/item/1005001592825712.html?_evo_buckets=165609%2C165598%2C188872%2C194277%2C224411%2C224373%2C176818&_t=gps-id%3ApcDetailBottomMoreThisSeller%2Cscm-url%3A1007.13339.169870.0%2Cpvid%3Ad8a27357-3570-41a7-8c62-fdf5f9a172c9%2Ctpp_buckets%3A21387%230%23233228%2310&gps-id=pcDetailBottomMoreThisSeller&pvid=d8a27357-3570-41a7-8c62-fdf5f9a172c9&scm=1007.13339.169870.0&scm-url=1007.13339.169870.0&scm_id=1007.13339.169870.0&sku_id=12000016713342689&spm=a2g2w.detail.0.0.36fe4ab0ve7lA4',
    'https://aliexpress.ru/item/1005003556155048.html?_evo_buckets=165609%2C165598%2C188871%2C194277%2C224411%2C224373%2C176818&_t=gps-id%3ApcDetailBottomMoreThisSeller%2Cscm-url%3A1007.13339.169870.0%2Cpvid%3Aa3e45df8-7ac5-4603-b472-8ade50e24cb0%2Ctpp_buckets%3A21387%230%23233228%238&gps-id=pcDetailBottomMoreThisSeller&item_id=1005003556155048&pvid=a3e45df8-7ac5-4603-b472-8ade50e24cb0&scm=1007.13339.169870.0&scm-url=1007.13339.169870.0&scm_id=1007.13339.169870.0&sku_id=12000026278135346&spm=a2g2w.detail.1000013.1.663d1e95JTc5BA'
]

fileName = 'ali.xlsx'

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='ali2', index=0)
count = 2
countRow = 2
excel_sheet['A1'] = 'id'
excel_sheet['B1'] = 'Название'
excel_sheet['C1'] = 'Цена'
excel_sheet['D1'] = 'Категория'
excel_sheet['E1'] = 'Описание'

for url in urls:
    countCol = 6
    countFoto = 1
    CountP = 1
    CountPn = 17
    CountPnZ = 18
    CountPZZ = 1
    largeFoto2 = []
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    name = soup.find('h1').get_text()
    foto = soup.find('div', class_='Product_GalleryBar__bar__156z6').find_all('img')
    price = soup.find('span',
                      class_='ali-kit_Base__base__1odrub ali-kit_Base__default__1odrub ali-kit_Base__strong__1odrub price ali-kit_Price__size-xl__12ybyf Product_Price__current__1uqb8 product-price-current').getText()

    descriptionRaw = soup.find('div', class_='ProductDescription-module_content__1xpeo')
    description = (str(descriptionRaw)).replace("\n", '')

    svoistva = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
                                                                                         class_='ali-kit_Base__base__1odrub ali-kit_Base__light__1odrub')

    svoistva_znach = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
                                                                                               class_='ali-kit_Base__base__1odrub ali-kit_Base__default__1odrub')

    print(description)
    print(price)
    print(name)

    # excel_sheet[f'A{count}'] = name
    excel_sheet[f'B{count}'] = name
    excel_sheet[f'C{count}'] = price
    excel_sheet[f'D{count}'] = "игрушки"
    excel_sheet[f'E{count}'] = description

    for f in foto:
        largeFoto = f['src']
        largeFoto2.append(largeFoto.replace('_50x50.jpg', ''))
    print(largeFoto2)

    for fot in largeFoto2:
        excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
        excel_sheet.cell(row=countRow, column=countCol).value = fot
        countFoto += 1
        countCol += 1

    datas = [i.get_text(strip=True) for i in svoistva]
    datas2 = [i.get_text() for i in svoistva_znach]

    for data in datas:
        excel_sheet.cell(row=1, column=CountPn).value = f'Свойство {CountP}'
        excel_sheet.cell(row=countRow, column=CountPn).value = data.replace(':', '')
        CountP += 1
        CountPn += 2

    for data2 in datas2:
        excel_sheet.cell(row=1, column=CountPnZ).value = f'Значение {CountPZZ}'
        excel_sheet.cell(row=countRow, column=CountPnZ).value = data2
        CountPnZ += 2
        CountPZZ += 1

    count += 1
    countRow += 1

excel_file.save(fileName)
