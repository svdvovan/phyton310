from os import mkdir

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from tkinter import *
import urllib3
import urllib.request

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

root = Tk()
root.geometry('850x250+1000+300')

l_url = Label(root, text="Введите url страницы категории:").grid(row=0, column=0, pady=10)
e_url = Entry(root, width=100)
e_url.grid(row=0, column=1, padx=15, sticky=E)

l_page = Label(root, text="Страницу с которой парсим:").grid(row=1, column=0, pady=10)
e_page = Entry(root, width=10)
e_page.insert(0, "1")
e_page.grid(row=1, column=1, padx=15, sticky=W)

l_lastPage = Label(root, text="Страницу по которую парсим:").grid(row=2, column=0, pady=10)
e_lastPage = Entry(root, width=10)
e_lastPage.insert(0, "10")
e_lastPage.grid(row=2, column=1, padx=15, sticky=W)


l_catName = Label(root, text="Название категории:").grid(row=3, column=0, pady=10)
e_catName = Entry(root, width=100)
# e_catName.insert(0, "1")
e_catName.grid(row=3, column=1, padx=15, sticky=W)

l_index = Label(root, text="Префикс файла и папки:").grid(row=4, column=0, pady=10)
e_index = Entry(root, width=20)
e_index.insert(0, "1")
e_index.grid(row=4, column=1, padx=15, sticky=W)


excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='genetrip2', index=0)

count = 2
countRow = 2
def createDir(catName, index):
    dirName = f'{index}_download_{catName}'
    mkdir(dirName)


def bigParser(e_url, page, lastPage, catName, index):
    global count
    global countRow
    for p in range(page, lastPage+1):
            if p==1:
                path2=e_url
            else:
                path2 = f'{e_url}?page={p}'
            print(path2)

            excel_sheet['A1'] = 'id'
            excel_sheet['B1'] = 'Название'
            excel_sheet['C1'] = 'Цена'
            excel_sheet['D1'] = 'Категория'
            excel_sheet['E1'] = 'Описание'
            excel_sheet['F1'] = 'Производитель'
            excel_sheet['G1'] = 'Главное фото'

            page = requests.get(path2, verify=False)
            soupCategory = BeautifulSoup(page.content, 'html.parser')
            allProductInCategory = soupCategory.find_all('div', class_='product-name')
            for a in allProductInCategory:
                for link in a.find_all('a'):
                    productPage=[]
                    productPage2 = (link.get('href'))
                    productPage.append(productPage2)
                    for p2 in productPage:
                        countCol = 8
                        countFoto = 1
                        CountP = 1
                        CountPn = 20
                        CountPnZ = 21
                        CountPZZ = 1
                        largeFoto2 = []

                        print(p2)
                        pageTovar = requests.get(p2, verify=False)
                        soup = BeautifulSoup(pageTovar.content, 'html.parser')

                        name = soup.find('h1', class_='h1-prod-name').getText()
                        try:

                            mainFoto = soup.find('a', id="zoom1").get('href')

                        except:
                            mainFoto =''
                        try:
                            foto = soup.find('div', class_='image-additional owl-carousel').find_all('a',
                                                                                                 class_='thumbnail')

                            for f in foto:
                                largeFoto = f['href']
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
                                nameFotoUrl = largeFoto.split("/")[-1]
                                req = urllib.request.Request(url=largeFoto, headers=headers)
                                urlImg = urllib.request.urlopen(req).read()
                                out = open(f'{index}_download_{catName}/{nameFotoUrl}', "wb")
                                out.write(urlImg)
                                out.close
                                largeFoto2.append(nameFotoUrl)
                            print(largeFoto2)

                            for fot in largeFoto2:
                                excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
                                excel_sheet.cell(row=countRow, column=countCol).value = fot
                                countFoto += 1
                                countCol += 1
                        except:
                            foto =''
                        id = soup.find(itemprop="model").getText()
                        price = soup.find('span',
                                          class_='autocalc-product-price').getText()
                        try:
                            manufac = soup.find(itemprop="brand").getText()
                        except:
                            manufac = ''

                        description1 = soup.find('div', id='tab-description')
                        description = str(description1)
                        try:
                            svoistva = soup.find('table', class_='table table-bordered').find_all('td',
                                                                                                  itemprop='name')

                            svoistva_znach = soup.find('table', class_='table table-bordered').find_all('td',
                                                                                                        itemprop='value')
                        except:
                            svoistva =''
                            svoistva_znach =''

                        print(id)
                        print(manufac)
                        print(description)
                        print(price)
                        print(name)
                        #
                        excel_sheet[f'A{count}'] = id
                        excel_sheet[f'B{count}'] = name
                        excel_sheet[f'C{count}'] = price
                        excel_sheet[f'D{count}'] = catName
                        excel_sheet[f'E{count}'] = description
                        excel_sheet[f'F{count}'] = manufac

                        # for f in foto:
                        #     largeFoto = f['href']
                        #     largeFoto2.append(largeFoto)
                        # print(largeFoto2)
                        #
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
                        nameFotoUrl = mainFoto.split("/")[-1]
                        req = urllib.request.Request(url=mainFoto, headers=headers)
                        urlImg = urllib.request.urlopen(req).read()
                        out = open(f'{index}_download_{catName}/{nameFotoUrl}', "wb")
                        out.write(urlImg)
                        out.close

                        excel_sheet[f'G{count}'] = nameFotoUrl

                        # for fot in largeFoto2:
                        #     excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
                        #     excel_sheet.cell(row=countRow, column=countCol).value = fot
                        #     countFoto += 1
                        #     countCol += 1

                        datas = [i.get_text(strip=True) for i in svoistva]
                        datas2 = [i.get_text() for i in svoistva_znach]

                        for data in datas:
                            excel_sheet.cell(row=1, column=CountPn).value = f'Свойство {CountP}'
                            excel_sheet.cell(row=countRow, column=CountPn).value = data.replace(':', '')
                            CountP += 1
                            CountPn += 2

                        for data2 in datas2:
                            excel_sheet.cell(row=1, column=CountPnZ).value = f'Значение {CountPZZ}'
                            excel_sheet.cell(row=countRow, column=CountPnZ).value = data2
                            CountPnZ += 2
                            CountPZZ += 1

                        count += 1
                        countRow += 1

                excel_file.save(f'{index}_{catName}.xlsx')


def getUrl():
    urls = e_url.get()
    catName = e_catName.get()
    index = e_index.get()
    page = int(e_page.get())
    lastPage = int(e_lastPage.get())
    createDir(catName, index)
    bigParser(urls, page, lastPage, catName, index)

def clearAll():
    e_url.delete(0, END)
    e_catName.delete(0, END)



l_start = Button(root, text="Парсить", command=getUrl).grid(row=5, column=0, padx=10, pady=10, sticky=W)
# l_clear = Button(root, text="Очистить", command=lambda: e_url.delete(0, END)).grid(row=5, column=1, padx=0, pady=10,
#                                                                                    sticky=W)
l_clear = Button(root, text="Очистить URL/Категорию", command=clearAll).grid(row=5, column=1, padx=0, pady=10,
                                                                                   sticky=W)
l_close = Button(root, text="Закрыть", command=lambda: root.destroy()).grid(row=5, column=1, padx=20, pady=10, sticky=E)
root.mainloop()
