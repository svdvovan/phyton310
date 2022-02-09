from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from tkinter import *

root = Tk()
root.geometry('800x150+1000+300')

fileName = 'med-konfitur.xlsx'

l_url = Label(root, text="Введите url страницы:").grid(row=0, column=0, pady=10)
e_url = Entry(root, width=100)
e_url.grid(row=0, column=1, padx=15, sticky=E)

l_cat = Label(root, text="Название категории").grid(row=1, column=0, pady=10)
e_cat = Entry(root, width=50)
e_cat.grid(row=1, column=1, padx=15, sticky=W)

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='med-konfitur', index=0)

count = 2
countRow = 2


def bigParser(e_url, e_cat):
    global count
    global countRow

    excel_sheet['A1'] = 'id'
    excel_sheet['B1'] = 'Название'
    excel_sheet['C1'] = 'Цена'
    excel_sheet['D1'] = 'Категория'
    excel_sheet['E1'] = 'Бренд'
    excel_sheet['F1'] = 'Описание'
    countCol = 7
    countFoto = 1
    CountP = 1
    CountPn = 17
    CountPnZ = 18
    CountPZZ = 1
    largeFoto2 = []

    for url in e_url:

        page = requests.get(url)
        category = ''.join(e_cat)
        soup = BeautifulSoup(page.content, 'html.parser')
        name = soup.find('h1').get_text()
        foto = soup.find('div', class_='slide').find_all('img')
        id = soup.find('div', class_='sec-product-attrs').find('div').find_all('span')[1].getText()
        brand = soup.find('div', class_='sec-product-attrs').find_all('div')[1].find_all('span')[1].getText()
        try:
            price = soup.find('div',
                              class_='price-old').getText()

        except:
            price = 100
        finally:
            descriptionRaw = soup.find('div', itemprop='description')
            description = (str(descriptionRaw)).replace("\n", '')

            # svoistva = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
            #                                                                                      class_='ali-kit_Base__base__1odrub ali-kit_Base__light__1odrub')
            #
            # svoistva_znach = soup.find('div', class_='Characteristics-styles_wrapper__2xHnh').find_all('span',
            #                                                                                            class_='ali-kit_Base__base__1odrub ali-kit_Base__default__1odrub')

            print(id)
            print(description)
            print(price)
            print(name)
            print(brand)
            print(category)

            excel_sheet[f'A{count}'] = id
            excel_sheet[f'B{count}'] = name
            excel_sheet[f'C{count}'] = price
            excel_sheet[f'D{count}'] = category
            excel_sheet[f'E{count}'] = brand
            excel_sheet[f'F{count}'] = description

            for f in foto:
                largeFoto = f['src']
                largeFoto2.append('https://www.med-konfitur.ru'+largeFoto)
            print(largeFoto2)

            for fot in largeFoto2:
                excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
                excel_sheet.cell(row=countRow, column=countCol).value = fot
                countFoto += 1
                countCol += 1
            #
            # datas = [i.get_text(strip=True) for i in svoistva]
            # datas2 = [i.get_text() for i in svoistva_znach]
            #
            # for data in datas:
            #     excel_sheet.cell(row=1, column=CountPn).value = f'Свойство {CountP}'
            #     excel_sheet.cell(row=countRow, column=CountPn).value = data.replace(':', '')
            #     CountP += 1
            #     CountPn += 2
            #
            # for data2 in datas2:
            #     excel_sheet.cell(row=1, column=CountPnZ).value = f'Значение {CountPZZ}'
            #     excel_sheet.cell(row=countRow, column=CountPnZ).value = data2
            #     CountPnZ += 2
            #     CountPZZ += 1

            count += 1
            countRow += 1

    excel_file.save(fileName)


def getUrl():
    print(e_url.get())
    urls = [e_url.get()]
    cat = [e_cat.get()]
    bigParser(urls, cat)


l_start = Button(root, text="Парсить", command=getUrl).grid(row=3, column=0, padx=10, pady=10, sticky=W)
l_clear = Button(root, text="Очистить URL", command=lambda: e_url.delete(0, END)).grid(row=3, column=1, padx=0, pady=10,
                                                                                   sticky=W)
l_clear = Button(root, text="Очистить категорию", command=lambda: e_cat.delete(0, END)).grid(row=3, column=1, padx=150, pady=10,
                                                                                   sticky=W)
l_close = Button(root, text="Закрыть", command=lambda: root.destroy()).grid(row=3, column=1, padx=20, pady=10, sticky=E)
root.mainloop()
