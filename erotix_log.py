import os
from os import mkdir
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from tkinter import *
from tkinter import filedialog
import urllib3
import urllib.request
import random
import json
from tkinter import ttk
import tkinter as tk

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
headers = {
    'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'}

root = Tk()
root.title("Парсер")
root.geometry('850x900+1000+300')
isDownload = IntVar()

l_url = Label(root, text="Введите url страницы категории:").grid(row=0, column=0, pady=10)
e_url = Entry(root, width=100)
e_url.grid(row=0, column=1, padx=15, sticky=E)

l_page = Label(root, text="Страницу с которой парсим:").grid(row=1, column=0, pady=10)
e_page = Entry(root, width=10)
e_page.insert(0, "1")
e_page.grid(row=1, column=1, padx=15, sticky=W)

l_lastPage = Label(root, text="Страницу по которую парсим:").grid(row=2, column=0, pady=10)
e_lastPage = Entry(root, width=10)
e_lastPage.insert(0, "10")
e_lastPage.grid(row=2, column=1, padx=15, sticky=W)

l_catName = Label(root, text="Название категории:").grid(row=3, column=0, pady=10)
e_catName = Entry(root, width=100)
# e_catName.insert(0, "1")
e_catName.grid(row=3, column=1, padx=15, sticky=W)

l_index = Label(root, text="Префикс файла:").grid(row=4, column=0, pady=10)
e_index = Entry(root, width=50)
e_index.insert(0, "1")
e_index.grid(row=4, column=1, padx=15, sticky=W)

l_dir = Label(root, text="Имя корневой папки").grid(row=5, column=0, pady=10)
e_dir = Entry(root, width=50)
e_dir.insert(0, "1_download")
e_dir.grid(row=5, column=1, padx=15, sticky=W)

l_isDownload_ = Label(root, text="Скачать фото?").grid(row=6, column=0, pady=10)
isDownload_checkbutton = Checkbutton(text="Скачать", variable=isDownload).grid(row=6, column=1, pady=10, sticky=W)

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='genetrip2', index=0)

log = tk.Text(root)
log.grid(row=11, column=1)
l_log = Label(root, text="Окно сообщения").grid(row=10, column=1, pady=10)
log.insert(INSERT, "Hello.....")

#TODO скрол не получился.....
# scrollbar = ttk.Scrollbar(log, orient='vertical', command=log.yview)
# scrollbar.grid(row=0, column=1, sticky='ns')
# log['yscrollcommand'] = scrollbar.set
# log.config(yscrollcommand=scrollbar.set)
# root_t = tk.Tk()
# text = tk.Text(log, height=8)
# text.grid(row=0, column=0, sticky='ew')
# scrollbar = ttk.Scrollbar(text, orient='vertical', command=text.yview)
# scrollbar.grid(row=0, column=1, sticky='ns')
# text['yscrollcommand'] = scrollbar.set
# root.grid_rowconfigure(1, weight=0)
# root.grid_columnconfigure(1, weight=0)
# textWidget = Entry(root, width=0)
# textWidget.insert(0, "1_download")

count = 2
countRow = 2


def createDir(catName, dir):
    dirName = f'{dir}_{catName}'
    try:
        mkdir(dirName)
    except:
        print('папка существует')

# ?page=2
# https://erotix.by/vibro-stimulyatory/?limit=10000
def bigParser(e_url, page, lastPage, catName, index, dir, isDownload2):
    global count
    global countRow
    for p in range(page, lastPage + 1):
        if p == 1:
            path2 = e_url
        else:
            path2 = f'{e_url}?page={p}'
        print(path2)

        excel_sheet['A1'] = 'id'
        excel_sheet['B1'] = 'Название'
        excel_sheet['C1'] = 'Цена'
        excel_sheet['D1'] = 'Категория'
        excel_sheet['E1'] = 'Описание'
        excel_sheet['F1'] = 'Производитель'
        excel_sheet['G1'] = 'Статус'

        page = requests.get(path2, headers=headers, verify=False)
        soupCategory = BeautifulSoup(page.content, 'html.parser')
        allProductInCategory = soupCategory.find_all('div', class_='h4')
        for a in allProductInCategory:
            for link in a.find_all('a'):
                productPage = []
                productPage2 = (link.get('href'))
                productPage.append(productPage2)
                for p2 in productPage:
                    countCol = 8
                    countFoto = 1
                    CountP = 1
                    CountPn = 20
                    CountPnZ = 21
                    CountPZZ = 1
                    largeFoto2 = []


                    pageTovar = requests.get(p2, headers=headers, verify=False)
                    soup = BeautifulSoup(pageTovar.content, 'html.parser')

                    name = soup.find('h1').getText()
                    print(name)
                    ##############################################################
                    # try:
                    #     mainFoto = soup.find('a', id="zoom1").get('href')
                    #     req = urllib.request.Request(url=mainFoto, headers=headers)
                    #     urlImg = urllib.request.urlopen(req).read()
                    #     if (isDownload2 == 1):
                    #         out = open(f'{dir}_{catName}/{nameFoto(mainFoto)}', "wb")
                    #         excel_sheet[f'G{count}'] = nameFoto(mainFoto)
                    #         out.write(urlImg)
                    #         out.close
                    #     else:
                    #         excel_sheet[f'G{count}'] = mainFoto
                    #
                    # except:
                    #     mainFoto = ''
                    ##################################################################
                    try:
                        foto = soup.find('ul', class_='thumbnails').find_all('a',
                                                                             class_='thumbnail')
                        for f in foto:
                            largeFoto = f['href']
                            req = urllib.request.Request(url=largeFoto, headers=headers)
                            urlImg = urllib.request.urlopen(req).read()

                            if (isDownload2 == 1):
                                out = open(f'{dir}_{catName}/{nameFoto(largeFoto)}', "wb")
                                out.write(urlImg)
                                out.close
                                largeFoto2.append(nameFoto(largeFoto))
                            else:
                                # nameFotoUrl = largeFoto.split("/")[-1]
                                largeFoto2.append(largeFoto)
                        print(largeFoto2)

                        for fot in largeFoto2:
                            excel_sheet.cell(row=1, column=countCol).value = f'фото{countFoto}'
                            excel_sheet.cell(row=countRow, column=countCol).value = fot
                            countFoto += 1
                            countCol += 1
                    except:
                        foto = ''
                    ##################################################################

                    price = soup.find_all('div', class_='h2')[1].getText()

                    try:
                        breadcrumb1 = (soup.find('ul', class_='breadcrumb').find_all('li')[-3].getText()).strip()
                        breadcrumb2 = (soup.find('ul', class_='breadcrumb').find_all('li')[-2].getText()).strip()
                        breadcrumb = breadcrumb1 + "->" + breadcrumb2
                        print("Категория " + breadcrumb)

                    except:
                        breadcrumb = breadcrumb2


                    del_id = 'Код товара: '
                    del_manufac = 'Производитель:'
                    try:

                        id_raw = soup.find('ul', class_='list-unstyled').find_all('li')[1].getText()
                        id = id_raw.replace(del_id, '')

                    except:
                        id = '?????'
                    try:
                        manufac_raw = soup.find('ul', class_='list-unstyled').find_all('li')[0].getText()
                        manufac = manufac_raw.replace(del_manufac, '')

                    except:
                        manufac = '?????'

                    try:
                        dostup = soup.find('div', id='content-aside').find_all('span')[0].getText()
                        if dostup!="Под заказ":
                            dostup = 'В продаже'
                    except:
                        dostup = 'В продаже'
                    excel_sheet[f'G{count}'] = dostup

                    description1 = soup.find('div', id='tab-description')
                    description_raw = str(description1)
                    description = description_raw.replace(
                        '<span style="font-weight: 700; font-family: Arial, sans-serif; font-size: 14px;">Импортёр в страны ЕАЭС </span><span style="font-family: Arial, sans-serif; font-size: 14px;">ООО "Кисс Экспо", УНП 691769478, 220089, г. Минск, пр. Дзержинского, д. 11, пом.802</span><br/></p></div>',
                        '')

                    try:
                        svoistva = soup.find('table', class_='table table-bordered').find_all('td',
                                                                                              itemprop='name')

                        svoistva_znach = soup.find('table', class_='table table-bordered').find_all('td',
                                                                                                    itemprop='value')
                    except:
                        svoistva = ''
                        svoistva_znach = ''

                    print(id)
                    print(manufac)
                    print(description)
                    print(price)
                    print(name)



                    excel_sheet[f'A{count}'] = id
                    excel_sheet[f'B{count}'] = name
                    excel_sheet[f'C{count}'] = price
                    # excel_sheet[f'D{count}'] = catName
                    excel_sheet[f'D{count}'] = breadcrumb
                    excel_sheet[f'E{count}'] = description
                    excel_sheet[f'F{count}'] = manufac

                    datas = [i.get_text(strip=True) for i in svoistva]
                    datas2 = [i.get_text() for i in svoistva_znach]

                    for data in datas:
                        excel_sheet.cell(row=1, column=CountPn).value = f'Свойство {CountP}'
                        excel_sheet.cell(row=countRow, column=CountPn).value = data.replace(':', '')
                        CountP += 1
                        CountPn += 2

                    for data2 in datas2:
                        excel_sheet.cell(row=1, column=CountPnZ).value = f'Значение {CountPZZ}'
                        excel_sheet.cell(row=countRow, column=CountPnZ).value = data2
                        CountPnZ += 2
                        CountPZZ += 1

                    count += 1
                    countRow += 1
                log.insert(END, f'\n Страница-{p} {name}')
                log.update()
            excel_file.save(f'{dir}_{catName}/{index}_{catName}.xlsx')


def getUrl():
    urls = e_url.get()
    catName = e_catName.get()
    index = e_index.get()
    page = int(e_page.get())
    lastPage = int(e_lastPage.get())
    dir = e_dir.get()
    createDir(catName, dir)
    isDownload2 = isDownload.get()
    bigParser(urls, page, lastPage, catName, index, dir, isDownload2)


def clearAll():
    e_url.delete(0, END)
    e_catName.delete(0, END)


def nameFoto(mainFoto):
    sizeFoto = mainFoto.split("/")[-1]
    if (len(sizeFoto)) > 10:
        mainFoto = f'{os.path.splitext(sizeFoto)[0][:10]}{random.randint(0, 1000)}{os.path.splitext(sizeFoto)[1]}'
        return mainFoto


def openConfigFile():
    openFile = filedialog.askopenfilename(title="Выбор файла конфигурации", defaultextension=".json",
                                          filetypes=(("JSON", "*.json"),))
    # with open(openFile, encoding='utf-8', errors='ignore') as json_file:
    with open(openFile, encoding = 'cp1251') as json_file:
        data = json.load(json_file)
        clear_all()
        print(data)
        for p in data['config']:
            e_catName.insert(1, (p['name']))
            e_url.insert(1, (p['url']))
            e_dir.insert(1, (p['dir']))
            e_index.insert(1, (p['index']))
            e_lastPage.insert(0, (p['lastPage']))
            e_page.insert(0, (p['page']))


def saveConfigFile():
    saveFile = filedialog.asksaveasfilename(defaultextension=".json",
                                            filetypes=(("JSON files", "*.json"),
                                                       ))
    data = {
        'config': [{
            'name': e_catName.get(),
            'url': e_url.get(),
            'dir': e_dir.get(),
            'index': e_index.get(),
            'lastPage': e_lastPage.get(),
            'page': e_page.get()
        }
        ]
    }
    with open(saveFile, "w") as write_file:
        json.dump(data, write_file, ensure_ascii=False)


def clear_all():
    e_page.delete('0', END)
    e_lastPage.delete('0', END)
    e_catName.delete('0', END)
    e_url.delete('0', END)
    e_dir.delete('0', END)
    e_index.delete('0', END)



l_start = Button(root, text="Парсить", command=getUrl).grid(row=8, column=0, padx=10, pady=10, sticky=W)
l_clear = Button(root, text="Очистить URL/Категорию", command=clearAll).grid(row=8, column=1, padx=0, pady=10,
                                                                             sticky=W)
l_close = Button(root, text="Закрыть", command=lambda: root.destroy()).grid(row=8, column=1, padx=20, pady=10, sticky=E)

l_open = Button(root, text="Загрузить конфу", command=lambda: openConfigFile()).grid(row=9, column=0, padx=10, pady=10,
                                                                                     sticky=W)
l_save = Button(root, text="Сохранить конфу", command=lambda: saveConfigFile()).grid(row=9, column=1, padx=10, pady=10,
                                                                                     sticky=W)
root.mainloop()

